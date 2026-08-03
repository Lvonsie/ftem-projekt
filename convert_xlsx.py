# -*- coding: utf-8 -*-
"""
Konvertiert ein FTEM-Tool Import-Excel (Sheet "detail") in unser JSON-Format.

Aufruf:  python3 convert_xlsx.py <datei.xlsx> <sportart-id>
Ergebnis: ftem_data_<sportart-id>.json im Projektordner

Regeln:
- Sheet "detail": Spalten = Topic-Key, Topic, Topic-Group, SubTopic-Key, SubTopic,
  SortKey, dann die 10 Stufen F1..M (Spalten 7-16).
- "Alterskategorie" wird NICHT als Thema uebernommen, sondern als "ages"
  (Beschriftung der Stufen-Koepfe) gespeichert.
- Gruppen-Zuordnung: training -> "Sport & Athlet:in";
  Material/equipment -> "Material"; alles andere -> "Strukturen & Umfeld".
- Benachbarte Stufen-Zellen mit identischem Text werden zu einem Segment
  zusammengefasst (zusammengefasste Spalten wie im FTEM-Tool).
- "-" oder leere Zellen gelten als leer.
- Links im Format [[Text|https://...]] werden aus dem Zelltext entfernt und
  als Dokument-Buttons (l-Liste) uebernommen.
"""
import sys, os, json, re
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
STAGES = ["F1","F2","F3","T1","T2","T3","T4","E1","E2","M"]
LINK_RE = re.compile(r'\[\[([^|\]]*)\|([^\]\n]*?)\]\]?')   # [[Text|URL]] (tolerant: auch nur eine schliessende Klammer)
REF_RE = re.compile(r'\[\[[^\]|]*\]\]')                     # [[interner-verweis]] ohne URL -> entfernen

def clean(v):
    if v is None: return ""
    s = str(v).replace("\r\n","\n").replace("\r","\n")
    s = re.sub(r'[ \t]+\n', '\n', s)          # Leerzeichen vor Zeilenumbruch
    s = re.sub(r'\n{3,}', '\n\n', s)          # max. eine Leerzeile
    s = s.strip()
    if s in ("-", "–", "—"): return ""
    return s

def group_of(topic_key, topic, group_key):
    if topic_key == "equipment" or topic.strip() == "Material":
        return "Material"
    # Persoenliche Themen gehoeren immer zu "Sport & Athlet:in", auch wenn sie im
    # Excel (z. B. Biathlon) unter health_and_development einsortiert sind.
    if topic_key in ("sleep", "nutrition", "psyche") or topic.strip() in ("Schlaf", "Ernährung", "Psyche"):
        return "Sport & Athlet:in"
    if group_key == "training":
        return "Sport & Athlet:in"
    return "Strukturen & Umfeld"

# Einheitliche Reihenfolge der Themen in "Sport & Athlet:in" (ueber alle Sportarten).
# Nicht aufgefuehrte Themen (z. B. SL/RS/Speed, Cross-Bloecke, Akrobatik, Technik)
# kommen danach in ihrer Original-Reihenfolge aus dem Excel.
def sport_prio(title):
    t = title.strip()
    tests = [
        t.startswith("Trainingsstunden"),
        t.startswith("Makroplanung"),
        t == "Ernährung",
        t == "Schlaf",
        t == "Psyche",
        t == "Ausdauer",
        t.startswith("Mobilität"),
        t.startswith("Kraft"),
        t.replace("&", "und").startswith("Schnelligkeit und Agilität"),
        t == "Schnelligkeit",
        t.startswith("Koordination"),          # Skispringen/NoKo: nach Schnelligkeit
        t == "Technik & Taktik Übergeordnet",
    ]
    for i, hit in enumerate(tests):
        if hit: return i
    if t == "Schiessen":                       # Biathlon: nach den Technik-Bloecken
        return len(tests) + 1
    return len(tests)

# Einheitliche Reihenfolge in "Strukturen & Umfeld"
def struct_prio(title):
    t = title.strip().replace(" und ", " & ")
    tests = [
        t == "Fördergefässe",
        t.startswith("Förderstrukturen"),
        t == "Selektionen",
        t == "Umfeldmanagement",
    ]
    for i, hit in enumerate(tests):
        if hit: return i
    return len(tests)

# Spalten des "homepage"-Sheets: 7-9 F1-F3, 11-14 T1-T4, 16-17 E1-E2, 19 M
HOME_COLS = [("F1",7),("F2",8),("F3",9),("T1",11),("T2",12),("T3",13),("T4",14),("E1",16),("E2",17),("M",19)]

def parse_cell(v):
    """Text + Links ([[Text|URL]]) aus einer Zelle loesen."""
    links = [{"text": m.group(1).strip() or "Dokument", "href": m.group(2).strip()}
             for m in LINK_RE.finditer(v) if m.group(2).strip().startswith("http")]
    txt = LINK_RE.sub('', v)
    txt = REF_RE.sub('', txt)
    txt = re.sub(r'[ \t]+\n', '\n', txt)
    txt = re.sub(r'\n{3,}', '\n\n', txt).strip()
    return txt, links

def parse_homepage(wb):
    """Sheet "homepage" -> Kurz-Zusammenfassung pro Stufe fuer die Startseiten-Popups.
    Zeile "Einleitung/introduction-home" liefert den Einleitungstext pro Phase (F1/T1/E1/M),
    die uebrigen Zeilen (WAS/WIE VIEL/Umfeld & Struktur) Texte pro Entwicklungsstufe."""
    name = next((s for s in wb.sheetnames if s.lower().strip() == "homepage"), None)
    if not name:
        return None
    ws = wb[name]
    intro = {}
    sections = []
    for r in range(2, ws.max_row + 1):
        sub = clean(ws.cell(r, 5).value)
        if not sub:
            continue
        key = clean(ws.cell(r, 4).value)
        cells = {}
        for st, c in HOME_COLS:
            txt, links = parse_cell(clean(ws.cell(r, c).value))
            if txt or links:
                cells[st] = {"v": txt, "l": links}
        if key == "introduction-home" or (sub == "Einleitung" and not intro):
            intro = {"f": cells.get("F1", {}).get("v", ""), "t": cells.get("T1", {}).get("v", ""),
                     "e": cells.get("E1", {}).get("v", ""), "m": cells.get("M", {}).get("v", "")}
        elif cells:
            sections.append({"title": sub, "cells": cells})
    if not intro and not sections:
        return None
    return {"intro": intro, "sections": sections}

def convert(xlsx_path, sport_id):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["detail"]
    themes = []          # in Reihenfolge des Excels
    by_title = {}
    ages = {}
    for r in range(2, ws.max_row + 1):
        topic = clean(ws.cell(r, 2).value)
        if not topic:
            continue
        topic_key = clean(ws.cell(r, 1).value)
        group_key = clean(ws.cell(r, 3).value)
        label = clean(ws.cell(r, 5).value)
        cells = [clean(ws.cell(r, c).value) for c in range(7, 17)]
        if topic == "Alterskategorie":
            for i, s in enumerate(STAGES):
                ages[s] = cells[i].replace("\n", " ").strip()
            continue
        if topic not in by_title:
            t = {"title": topic, "group": group_of(topic_key, topic, group_key), "rows": []}
            by_title[topic] = t
            themes.append(t)
        # Links [[Text|URL]] aus dem Zelltext loesen
        parsed = []
        for v in cells:
            links = [{"text": m.group(1).strip() or "Dokument", "href": m.group(2).strip()}
                     for m in LINK_RE.finditer(v)
                     if m.group(2).strip().startswith("http")]
            txt = LINK_RE.sub('', v)
            txt = REF_RE.sub('', txt)
            txt = re.sub(r'[ \t]+\n', '\n', txt)
            txt = re.sub(r'\n{3,}', '\n\n', txt).strip()
            parsed.append((txt, links))
        # Segmente: benachbarte identische Zellen (Text UND Links) zusammenfassen,
        # aber NUR innerhalb derselben Phase (F, T, E, M) - nie phasenuebergreifend.
        def phase_of(i):
            return "F" if i < 3 else "T" if i < 7 else "E" if i < 9 else "M"
        segs = []
        for i, (v, links) in enumerate(parsed):
            if (segs and segs[-1]["v"] == v and segs[-1]["l"] == links
                    and phase_of(segs[-1]["from"]) == phase_of(i)):
                segs[-1]["to"] = i
            else:
                segs.append({"v": v, "from": i, "to": i, "l": links})
        by_title[topic]["rows"].append({"label": label or None, "segs": segs})
    # Einheitliche Themen-Reihenfolge: Sport & Athlet:in sortiert, dann Material,
    # dann Strukturen & Umfeld sortiert (stabil: Unbekanntes behaelt Excel-Reihenfolge)
    sa = [t for t in themes if t["group"] == "Sport & Athlet:in"]
    ma = [t for t in themes if t["group"] == "Material"]
    su = [t for t in themes if t["group"] == "Strukturen & Umfeld"]
    rest = [t for t in themes if t["group"] not in ("Sport & Athlet:in", "Material", "Strukturen & Umfeld")]
    sa.sort(key=lambda t: sport_prio(t["title"]))
    su.sort(key=lambda t: struct_prio(t["title"]))
    themes = sa + ma + su + rest

    data = {"stages": STAGES, "ages": ages, "themes": themes}
    home = parse_homepage(wb)
    if home:
        data["home"] = home
    out = os.path.join(BASE, "ftem_data_" + sport_id + ".json")
    json.dump(data, open(out, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    n_rows = sum(len(t["rows"]) for t in themes)
    print("written", os.path.basename(out), "| themes:", len(themes), "| rows:", n_rows,
          "| ages:", bool(ages), "| home:", bool(home))

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
